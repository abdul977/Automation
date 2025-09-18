import sys
import json

def excel_to_json(file_path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl is required. Please install it with: py -m pip install openpyxl")
        return None

    # Load the Excel file
    workbook = load_workbook(file_path, data_only=True)

    # Get the first sheet
    sheet = workbook.active

    # Get all data from the sheet
    data = []

    # Get headers from the first row
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value if cell.value is not None else "")

    # Get data from remaining rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for i, value in enumerate(row):
            if i < len(headers):
                row_data[headers[i]] = value
        data.append(row_data)

    # Convert to JSON
    json_data = json.dumps(data, default=str, indent=2)
    return json_data

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python script.py <excel_file_path>")
    else:
        file_path = sys.argv[1]
        try:
            result = excel_to_json(file_path)
            if result:
                print(result)
        except Exception as e:
            print(f"Error: {e}")
